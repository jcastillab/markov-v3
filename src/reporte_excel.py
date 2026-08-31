"""Libro Excel de auditoria y visualizacion de modelos de pronostico."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from canonical import load_config
except ModuleNotFoundError:
    from src.canonical import load_config


GREEN, YELLOW, RED, BLUE, DARK = "C6EFCE", "FFEB9C", "FFC7CE", "D9EAF7", "1F4E78"

PREDICTION_FILES = {
    "E00_M3_BASE": "outputs/predictions/E00_M3_BASE.csv",
    "E01_M3_INGRESO_CALIBRADO": "outputs/predictions/E01_M3_INGRESO_CALIBRADO.csv",
    "E02_M3_P32_RAW": "outputs/predictions/E02_M3_P32_RAW.csv",
    "E03_M3_P32_CANONICO": "outputs/predictions/E03_M3_P32_CANONICO.csv",
    "E04_M3_P32_REG": "outputs/predictions/E04_M3_P32_REG.csv",
    "E05_SEMIMARKOV_P32_RC": "outputs/predictions/E05_SEMIMARKOV_P32_RC.csv",
    "E06_SEMIMARKOV_P32_RC_SS": "outputs/predictions/E06_SEMIMARKOV_P32_RC_SS.csv",
    "E07_SEMIMARKOV_P32_ALL": "outputs/predictions/E07_SEMIMARKOV_P32_ALL.csv",
    "RF_DIARIO_POOLED_FENO": "outputs/evaluation/predictions_feno.csv",
    "RF_DIARIO_POOLED_FENO_PODA": "outputs/evaluation/predictions_feno_poda.csv",
    "RF_DIARIO_POOLED_FENO_CLIMA": "outputs/evaluation/predictions_feno_clima.csv",
    "RF_DIARIO_POOLED_FENO_PODA_CLIMA": "outputs/evaluation/predictions_feno_poda_clima.csv",
    "RF_RESIDUAL_M3_FENO": "outputs/evaluation/predictions_rf_residual_m3_feno.csv",
    "RF_H1_H7_FENO": "outputs/evaluation/predictions_rf_h1_h7_feno.csv",
    "GLM_NB_FENO": "outputs/evaluation/predictions_glm_nb_feno.csv",
    "GLM_NB_FENO_PODA": "outputs/evaluation/predictions_glm_nb_feno_poda.csv",
    "GLM_NB_FENO_CLIMA": "outputs/evaluation/predictions_glm_nb_feno_clima.csv",
    "GLM_NB_FENO_PODA_CLIMA": "outputs/evaluation/predictions_glm_nb_feno_poda_clima.csv",
    "NB_JERARQUICO": "outputs/evaluation/predictions_nb_jerarquico.csv",
    "M3_DIRICHLET_MULTINOMIAL": "outputs/evaluation/predictions_m3_dirichlet.csv",
}


def weekly_status(ratio: float | None) -> str:
    if ratio is None or not np.isfinite(ratio):
        return "SIN REAL / N.A."
    if 0.93 <= ratio <= 1.07:
        return "ACIERTO"
    if 0.90 <= ratio < 0.93 or 1.07 < ratio <= 1.10:
        return "CERCA"
    return "NO ACIERTO"


def _metrics(root: Path) -> pd.DataFrame:
    files = sorted((root / "outputs/evaluation").glob("metrics_*.csv"))
    frames = []
    for path in files:
        frame = pd.read_csv(path)
        frame["archivo_metricas"] = path.name
        frames.append(frame)
    result = pd.concat(frames, ignore_index=True, sort=False)
    result["causal"] = result["causal"].astype(str).str.lower().eq("true")
    result["comparacion_primaria"] = result["causal"] & result["n"].eq(714) & result["split"].eq("VALIDATION")
    return result


def _traces(root: Path) -> dict[str, pd.DataFrame]:
    result = {}
    for model, relative in PREDICTION_FILES.items():
        path = root / relative
        if not path.exists():
            continue
        frame = pd.read_csv(path)
        pred_col = "pred_bloque" if "pred_bloque" in frame else "pred"
        real_col = "real"
        if real_col not in frame or pred_col not in frame:
            continue
        frame = frame.rename(columns={pred_col: "proyectado", real_col: "real"})
        frame["modelo"] = model
        for col in ("fecha_origen", "fecha_objetivo"):
            if col in frame:
                frame[col] = pd.to_datetime(frame[col]).dt.date
        result[model] = frame
    return result


def _model_note(model: str) -> tuple[str, str, str]:
    if model.startswith("E00") or model.startswith("E01"):
        return ("M3 Markov", "Propaga RC, SS y AP con matriz por finca/periodo; ingreso RC fijo.", "Causal, interpretable; no representa edad dentro del estado.")
    if "P32" in model or "SEMIMARKOV" in model:
        return ("P32 / Semi Markov", "Usa seguimiento P32 y, en Semi Markov, hazards por edad.", "Retrospectivo no causal: P32 se levantó después de las ventanas.")
    if model.startswith("M3_PODA"):
        return ("M3 + podas", "Modifica ingreso RC o transición con poda/alineamiento rezagado 8-12 semanas.", "Causal; la señal no mejoró M3 en el backtest.")
    if model.startswith("M3_CLIMA"):
        return ("M3 + clima", "Modifica ingreso o transición con GDD causal de estación exterior 746.", "Solo LA PRADERA; clima exterior es proxy, no microclima.")
    if model.startswith("RF_"):
        return ("Random Forest", "Predice corte directo con conteos, historia causal, escala y horizonte; variantes añaden poda/clima.", "Buen WAPE; menor interpretabilidad y requiere vigilancia de drift.")
    if model.startswith("GLM_NB"):
        return ("GLM Negative Binomial", "Regresión directa NB2 con enlace log y features causales.", "Control interpretable; inestable con escala/colinealidad actual.")
    if model == "NB_JERARQUICO":
        return ("NB jerárquico", "Pooling Gamma-Poisson por finca, bloque y horizonte hacia media global.", "Mejora WAPE; cobertura predictiva insuficiente.")
    if model == "M3_DIRICHLET_MULTINOMIAL":
        return ("M3 Bayes Dirichlet", "Muestrea matrices posteriores Dirichlet centradas en M3 causal.", "Regulariza transiciones; no mejora WAPE ni cobertura aún.")
    return ("Experimental", "Ver hoja PARAMETROS.", "Revisar alcance y causalidad.")


def _write_table(ws, frame: pd.DataFrame, title: str, percent_columns=()):
    ws.append([title])
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(frame.columns)))
    ws.append(list(frame.columns))
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
    for row in frame.itertuples(index=False, name=None):
        ws.append(list(row))
    ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:{get_column_letter(len(frame.columns))}{len(frame)+2}"
    for i, col in enumerate(frame.columns, 1):
        width = min(48, max(12, len(str(col)) + 2, *(len(str(v)) + 1 for v in frame.iloc[:100, i - 1].fillna(""))))
        ws.column_dimensions[get_column_letter(i)].width = width
        if col in percent_columns:
            for cell in ws[get_column_letter(i)][2:]:
                cell.number_format = "0.00%"
    return len(frame) + 2


def _add_semaphores(ws, row_end: int, headers: list[str]):
    lookup = {name: i + 1 for i, name in enumerate(headers)}
    for label, threshold in (("wape", 0.30), ("bias_pct", 0.10)):
        if label in lookup:
            col = get_column_letter(lookup[label]); rng = f"{col}3:{col}{row_end}"
            if label == "bias_pct":
                ws.conditional_formatting.add(rng, ColorScaleRule(start_type="min", start_color=GREEN,
                    mid_type="percentile", mid_value=50, mid_color=YELLOW, end_type="max", end_color=RED))
            else:
                ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=[str(threshold)], fill=PatternFill("solid", fgColor=GREEN)))
                ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(threshold), "0.55"], fill=PatternFill("solid", fgColor=YELLOW)))
                ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0.55"], fill=PatternFill("solid", fgColor=RED)))
    for label in ("mae", "rmse"):
        if label in lookup:
            col = get_column_letter(lookup[label]); ws.conditional_formatting.add(
                f"{col}3:{col}{row_end}", ColorScaleRule(start_type="min", start_color=GREEN,
                mid_type="percentile", mid_value=50, mid_color=YELLOW, end_type="max", end_color=RED))
    if "acierto_pct" in lookup:
        col = get_column_letter(lookup["acierto_pct"]); rng = f"{col}3:{col}{row_end}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.93", "1.07"], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.90", "0.93"], fill=PatternFill("solid", fgColor=YELLOW)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["1.07", "1.10"], fill=PatternFill("solid", fgColor=YELLOW)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.90"], fill=PatternFill("solid", fgColor=RED)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["1.10"], fill=PatternFill("solid", fgColor=RED)))


def build_workbook(root: Path) -> Path:
    cfg, metrics, traces = load_config(root / "config/pipeline.yaml"), _metrics(root), _traces(root)
    acierto_by_model = {}
    for model, frame in traces.items():
        acierto_by_model[model] = np.where(frame.real != 0,
            1 - (frame.real - frame.proyectado) / frame.real, np.nan).mean()
    reports = root / cfg["paths"]["outputs"] / "reports"; reports.mkdir(parents=True, exist_ok=True)
    output = reports / "tablero_rendimiento_modelos.xlsx"
    wb = Workbook(); intro = wb.active; intro.title = "INICIO"
    intro.append(["Tablero de rendimiento - Pronóstico Freedom"])
    intro["A1"].font = Font(bold=True, size=18, color="FFFFFF"); intro["A1"].fill = PatternFill("solid", fgColor=DARK)
    intro.merge_cells("A1:F1")
    intro_rows = [
        ("Uso", "Filtre las hojas METRICAS, SEMANAL y DIARIO por modelo, finca, bloque y fecha."),
        ("Ranking primario", "Solo VALIDATION causal diaria n=714. No mezcla P32 retrospectivo, clima parcial ni métricas semanales."),
        ("Acierto", "Se calcula como 1 - (real - proyectado) / real, tanto diario como semanal. Para real=0 queda N.A.; no se deriva de WAPE."),
        ("Semáforo semanal", "ACIERTO: precisión relativa entre 93%-107%; CERCA: 90%-92,9% o 107,1%-110%; NO ACIERTO: fuera de rangos."),
        ("Semáforo de error", "WAPE <=30% verde; >30%-55% amarillo; >55% rojo. MAE/RMSE usan escala relativa de la tabla: menor es verde."),
        ("Acierto global", "La columna heredada acierto_global equivale a 1-WAPE; no es el indicador semanal de razón proyectado/real."),
        ("Cobertura", "SIN TRAZABILIDAD indica que la fase dejó métricas agregadas sin pronósticos por fecha; no se infiere cobertura."),
    ]
    for row in intro_rows: intro.append(row)
    intro.column_dimensions["A"].width = 24; intro.column_dimensions["B"].width = 115
    for cell in intro["A"][1:]: cell.font = Font(bold=True)

    primary = metrics[metrics.comparacion_primaria].sort_values("wape").drop_duplicates("experiment_id").copy()
    primary["wape_pct"] = primary.wape; primary["acierto_pct"] = primary.experiment_id.map(acierto_by_model)
    summary_cols = ["experiment_id", "wape_pct", "acierto_pct", "mae", "rmse", "bias_pct", "n"]
    ws = wb.create_sheet("RESUMEN")
    end = _write_table(ws, primary[summary_cols], "Ranking primario: VALIDATION causal diaria n=714", {"wape_pct", "acierto_global_pct", "bias_pct"})
    ws.conditional_formatting.add(f"B3:B{end}", ColorScaleRule(start_type="min", start_color=GREEN, mid_type="percentile", mid_value=50, mid_color=YELLOW, end_type="max", end_color=RED))
    _add_semaphores(ws, end, list(primary[summary_cols].columns))
    chart = BarChart(); chart.title = "WAPE de modelos comparables"; chart.y_axis.title = "WAPE"; chart.height = 9; chart.width = 18
    chart.add_data(Reference(ws, min_col=2, min_row=2, max_row=min(end, 12)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=min(end, 12))); ws.add_chart(chart, "J3")

    ws = wb.create_sheet("METRICAS")
    detail = metrics.copy(); detail["wape_pct"] = detail.wape; detail["acierto_global_pct"] = detail.acierto_global
    end = _write_table(ws, detail, "Todas las métricas: revisar causalidad, población y escala antes de comparar", {"wape", "wape_pct", "acierto_global", "acierto_global_pct", "bias_pct", "coverage_interval_80", "coverage_interval_95"})
    _add_semaphores(ws, end, list(detail.columns))

    expected = set()
    if "E00_M3_BASE" in traces:
        base = traces["E00_M3_BASE"]
        expected = set(map(tuple, base[["finca", "bloque", "fecha_origen"]].drop_duplicates().itertuples(index=False, name=None)))
    coverage_rows, weekly_frames, daily_frames = [], [], []
    for model, frame in traces.items():
        if not {"finca", "bloque", "fecha_origen", "fecha_objetivo"}.issubset(frame.columns):
            continue
        daily = frame.copy(); daily["error"] = daily.proyectado - daily.real
        daily["error_abs"] = daily.error.abs(); daily["ratio_proyeccion_pct"] = np.where(daily.real != 0, daily.proyectado / daily.real, np.nan)
        daily["acierto_pct"] = np.where(daily.real != 0, 1 - (daily.real - daily.proyectado) / daily.real, np.nan)
        daily_frames.append(daily)
        weekly = daily.groupby(["modelo", "finca", "fecha_origen"], as_index=False).agg(
            real_semana=("real", "sum"), proyectado_semana=("proyectado", "sum"), dias_pronosticados=("fecha_objetivo", "nunique"))
        weekly["ratio_proyeccion_pct"] = np.where(weekly.real_semana != 0, weekly.proyectado_semana / weekly.real_semana, np.nan)
        weekly["acierto_pct"] = np.where(weekly.real_semana != 0, 1 - (weekly.real_semana - weekly.proyectado_semana) / weekly.real_semana, np.nan)
        weekly["indicador_semanal"] = weekly.ratio_proyeccion_pct.map(weekly_status)
        weekly_frames.append(weekly)
        for finca, finca_weekly in weekly.groupby("finca"):
            expected_finca = {key[2] for key in expected if key[0] == finca}
            found = set(finca_weekly.fecha_origen)
            coverage_rows.append({"modelo": model, "finca": finca, "semanas_esperadas": len(expected_finca), "semanas_con_pronostico": len(found),
                              "semanas_sin_pronostico": len(expected_finca - found),
                              "aciertos_verdes": int(finca_weekly.indicador_semanal.eq("ACIERTO").sum()),
                              "cerca_amarillos": int(finca_weekly.indicador_semanal.eq("CERCA").sum()),
                              "no_aciertos_rojos": int(finca_weekly.indicador_semanal.eq("NO ACIERTO").sum()),
                              "estado_trazabilidad": "TRAZABLE"})
    covered = {r["modelo"] for r in coverage_rows}
    for model in metrics.experiment_id.drop_duplicates():
        if model not in covered:
            coverage_rows.append({"modelo": model, "finca": "TODAS", "semanas_esperadas": np.nan, "semanas_con_pronostico": np.nan,
                                  "semanas_sin_pronostico": np.nan, "aciertos_verdes": np.nan, "cerca_amarillos": np.nan,
                                  "no_aciertos_rojos": np.nan, "estado_trazabilidad": "SIN TRAZABILIDAD"})
    coverage = pd.DataFrame(coverage_rows).sort_values("modelo")
    ws = wb.create_sheet("COBERTURA")
    end = _write_table(ws, coverage, "Cobertura y acierto semanal por modelo")
    status_col = get_column_letter(list(coverage.columns).index("aciertos_verdes") + 1)
    ws.conditional_formatting.add(f"{status_col}3:{status_col}{end}", ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=YELLOW, end_type="max", end_color=GREEN))

    weekly_all = pd.concat(weekly_frames, ignore_index=True).sort_values(["modelo", "fecha_origen"])
    ws = wb.create_sheet("SEMANAL")
    end = _write_table(ws, weekly_all, "Proyectado vs real por semana de origen", {"ratio_proyeccion_pct", "acierto_pct"})
    col = get_column_letter(list(weekly_all.columns).index("indicador_semanal") + 1)
    ws.conditional_formatting.add(f"{col}3:{col}{end}", CellIsRule(operator="equal", formula=['"ACIERTO"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"{col}3:{col}{end}", CellIsRule(operator="equal", formula=['"CERCA"'], fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add(f"{col}3:{col}{end}", CellIsRule(operator="equal", formula=['"NO ACIERTO"'], fill=PatternFill("solid", fgColor=RED)))
    _add_semaphores(ws, end, list(weekly_all.columns))

    daily_all = pd.concat(daily_frames, ignore_index=True).sort_values(["modelo", "fecha_objetivo"])
    ws = wb.create_sheet("DIARIO")
    end = _write_table(ws, daily_all, "Proyectado vs real diario", {"ratio_proyeccion_pct", "acierto_pct"})
    _add_semaphores(ws, end, list(daily_all.columns))

    control_rows = []
    for model, frame in traces.items():
        if not {"real", "proyectado"}.issubset(frame.columns):
            continue
        y, p = frame.real.to_numpy(float), frame.proyectado.to_numpy(float)
        denom = np.abs(y).sum()
        calculated = {"wape_recalculado": np.abs(p - y).sum() / denom if denom else np.nan,
                      "mae_recalculado": np.abs(p - y).mean(),
                      "rmse_recalculado": np.sqrt(np.mean((p - y) ** 2)),
                      "acierto_relativo_medio": np.where(y != 0, 1 - (y - p) / y, np.nan).mean()}
        source = metrics[metrics.experiment_id.eq(model)]
        control_rows.append({"modelo": model, "n_predicciones": len(frame),
                             **calculated, "metricas_fuente": "DISPONIBLE" if len(source) else "N.A."})
    ws = wb.create_sheet("CONTROL_RF_GLM")
    end = _write_table(ws, pd.DataFrame(control_rows), "Control: métricas recalculadas desde predicciones trazables", {"wape_recalculado", "acierto_relativo_medio"})
    _add_semaphores(ws, end, list(pd.DataFrame(control_rows).columns))

    bayes_cols = [c for c in ("experiment_id", "wape", "coverage_interval_80", "coverage_interval_95", "ancho_medio_intervalo", "n") if c in metrics]
    ws = wb.create_sheet("BAYES")
    bayes = metrics[metrics.experiment_id.isin(["NB_JERARQUICO", "M3_DIRICHLET_MULTINOMIAL"])][bayes_cols]
    _write_table(ws, bayes, "Métricas bayesianas e incertidumbre", {"wape", "coverage_interval_80", "coverage_interval_95"})

    parameter_rows = []
    for section in ("m3", "pruning", "climate", "random_forest", "supervised", "bayes", "semimarkov"):
        for key, value in cfg.get(section, {}).items(): parameter_rows.append({"componente": section, "parametro": key, "valor": json.dumps(value, ensure_ascii=False, default=str)})
    ws = wb.create_sheet("PARAMETROS"); _write_table(ws, pd.DataFrame(parameter_rows), "Parámetros vigentes: config/pipeline.yaml")

    model_rows = []
    for model in metrics.experiment_id.drop_duplicates():
        family, mechanism, caution = _model_note(model)
        model_rows.append({"modelo": model, "familia": family, "como_pronostica": mechanism, "ventajas_y_limites": caution})
    ws = wb.create_sheet("MODELOS"); _write_table(ws, pd.DataFrame(model_rows), "Ficha funcional por modelo")
    wb.save(output)
    return output


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    print(build_workbook(root))
